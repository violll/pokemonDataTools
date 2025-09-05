import argparse
import json
import os
import re
import sys

import helper
import numpy as np
import openpyxl
import pandas as pd
import yaml

# adds the path absolutely so the code can be run from anywhere
sys.path.insert(0, helper.getAbsPath(__file__, 2))

from src.pokeapi import main as pokeapi
from src.utils.google_sheets_api import google_sheets_api


class GroupData:
    def __init__(self, routes=[], encounters=[], assign_me={}) -> None:
        self.routes = routes
        self.encounters = encounters
        self.assign_me = assign_me

        self.flags = {"Honey": False}

    def __repr__(self) -> str:
        return json.dumps(self.assign_me, indent=4)


class LocalGame:
    def __init__(self, file_path) -> None:
        # open NRotM spreadsheet
        self.wb = openpyxl.load_workbook(filename=file_path)

        self.ws = self.wb["Team & Encounters"]

        self.route_col = self.get_col("Location")
        self.encounter_col = self.get_col("Encounter")
        self.route_order = [
            route.value
            for route in self.ws[self.route_col]
            if route.value not in [None, "Location"]
        ]
        self.game_name = self.wb["Tracker"]["A1"].value

        # init pokeapi calls
        self.PokeAPI = pokeapi.PokeapiAccess()
        self.pokedex = self.PokeAPI.get_pokedex_from_region(self.game_name)
        self.evo_lines = self.PokeAPI.get_evo_lines_from_pokedex(self.pokedex)

        # initialize encounter info
        self.encounter_data = self.get_encounter_data()
        self.encounter_table = self.get_encounter_table()

        # game specific information
        if self.game_name == "Platinum":
            self.honey_mons = [
                "Aipom",
                "Heracross",
                "Wurmple",
                "Burmy",
                "Combee",
                "Cherubi",
                "Munchlax",
            ]

    def get_col(self, value):
        ws = self.wb["Team & Encounters"]
        for row in ws.iter_rows():
            for col in row:
                if col.value == value:
                    return col.column_letter

        return None

    def get_encounter_data(self):
        res = []

        dv_list = self.ws.data_validations.dataValidation
        for dv in dv_list:
            cells = [
                str(c) for c in list(dv.sqref.ranges) if self.encounter_col in str(c)
            ]  # a list of CellRange elements

            if cells == []:
                continue

            encounters = dv.formula1.replace('"', "").split(",")  # a list of encounters

            cells = self.expand_colon(cells)
            for c in cells:
                loc = c.replace(self.encounter_col, chr(ord(self.encounter_col) - 1))
                route = self.ws[loc].value
                res.extend([route, enc, 1] for enc in encounters)

        return res

    def expand_colon(self, cells):
        res = []
        for cell in cells:
            if ":" not in cell:
                res.append(cell)
            else:
                start, stop = map(int, cell.replace(self.encounter_col, "").split(":"))
                while start <= stop:
                    res.append(self.encounter_col + str(start))
                    start += 1

        return res

    def get_encounter_table(self):
        # initial df without manipulation
        df = pd.DataFrame.from_records(
            self.encounter_data, columns=["Route", "Encounter", "Value"]
        ).pivot(index="Route", columns="Encounter", values="Value")

        return self.consolidate_evo_lines(
            df.reindex([route for route in self.route_order if route in list(df.index)])
        ).dropna(axis=0, how="all")

    def consolidate_evo_lines(self, df):
        res = []

        for line in self.evo_lines:
            relevant_line = [
                re.search(
                    r"^{}".format(mon), "\n".join(df.columns), flags=re.I | re.M
                ).group()
                for mon in line
                if re.search(
                    r"^{}".format(mon), "\n".join(df.columns), flags=re.I | re.M
                )
            ]
            if len(relevant_line) != 0:
                df_filtered = pd.DataFrame(df[relevant_line].agg("max", axis="columns"))
                df_filtered.columns = ["/".join(relevant_line)]
                res.append(df_filtered)

        return pd.concat(res, axis=1)


class CloudGame(LocalGame):
    def __init__(self, run_config):
        self.run_config = run_config

        # read game name
        api_call_params = {
            "spreadsheetId": self.run_config["spreadsheet_id"],
            "range": "Tracker!A1",
        }

        self.api = google_sheets_api.GoogleSheetsApi(
            self.run_config["creds_path"], self.run_config["token_path"]
        )
        self.game_name = self.api.main(api_call_params, "ss_values")["values"][0][0]

        # check if NRotM spreadsheet has been parsed yet
        if not os.path.exists(self.run_config["output_json_path"]):
            api_call_params = {
                "spreadsheetId": self.run_config["spreadsheet_id"],
                "ranges": self.run_config["spreadsheet_range"],
                "fields": "sheets.data.rowData.values.dataValidation,sheets.data.rowData.values.userEnteredValue.stringValue",
            }

            self.api.main(api_call_params, "ss", self.run_config["output_json_path"])

        with open(self.run_config["output_json_path"], "r") as f:
            self.sheet_data = json.load(f)

        # init pokeapi calls
        self.pokeapi = pokeapi.PokeapiAccess()
        self.pokedex = self.pokeapi.get_pokedex_from_region(self.game_name)
        self.evo_lines = self.pokeapi.get_evo_lines_from_pokedex(self.pokedex)

        # get encounter status
        n_routes = len(self.sheet_data["sheets"][0]["data"][0]["rowData"])

        api_call_params = {
            "spreadsheetId": self.run_config["spreadsheet_id"],
            "range": f"Team & Encounters!N5:N{5 + n_routes}",
        }
        self.encounter_status = self.api.main(api_call_params, "ss_values")["values"]

        # initialize encounter info
        self.encounter_data, self.encounters, self.route_order, self.failed_routes = (
            self.get_encounter_data()
        )
        self.encounter_table = self.get_encounter_table()

        # game specific information
        if self.game_name == "Platinum":
            self.honey_mons = [
                "Aipom",
                "Heracross",
                "Wurmple",
                "Burmy",
                "Combee",
                "Cherubi",
                "Munchlax",
            ]

    def get_encounter_data(self):
        possible_encounters = []
        received_encounters = {}
        route_order = []
        failed_routes = []

        for i in range(len(self.encounter_status)):
            encounter_status = self.encounter_status[i]
            row = self.sheet_data["sheets"][0]["data"][0]["rowData"][i]

            route_data, encounter_data = row["values"]

            route_name = route_data["userEnteredValue"]["stringValue"]
            encounter_name = encounter_data.get("userEnteredValue", {}).get(
                "stringValue", ""
            )

            route_order.append(route_name)

            # add encounter to the dict
            if encounter_name != "" and (
                encounter_status != [] and encounter_status[0] != "F"
            ):
                received_encounters[route_name] = encounter_name

            if encounter_status != [] and encounter_status[0] == "F":
                failed_routes.append(route_name)

            # get possible encounters for the route
            possible_route_encounters = encounter_data["dataValidation"]["condition"][
                "values"
            ]

            possible_encounters.extend(
                (route_name, mon["userEnteredValue"], 1)
                for mon in possible_route_encounters
            )
        return possible_encounters, received_encounters, route_order, failed_routes


class NRotMEncounterRouting:
    def __init__(self) -> None:
        # user can include an optional argument to check for a json file of encounters
        # config file is required
        self.parser = self.init_parser()
        self.args = self.parser.parse_args()

        # read config file
        with open(self.args.config, "r") as f:
            self.run_config = yaml.safe_load(f)

        # initialize assigned encounters
        self.assigned_encounters = {}
        self.assigned_encounters_slice = []

        # get initial dataset of routes and encounters
        if self.args.cloud:
            self.GameData = CloudGame(self.run_config)

            # initialize slice history of encounter assignment
            self.encounter_tables = [self.GameData.encounter_table]

            # update assigned encounters
            group_data = GroupData(
                assign_me=self.GameData.encounters,
                routes=list(self.GameData.encounters.keys())
                + self.GameData.failed_routes,
                encounters=list(self.GameData.encounters.values()),
            )
            self.update(group_data, self.GameData.encounter_table, is_json=True)

        else:
            self.GameData = LocalGame(file_path=self.run_config["sheet_path"])

            # initialize slice history of encounter assignment
            self.encounter_tables = [self.GameData.encounter_table]

            # update assigned encounters if file exists
            if os.path.exists(self.run_config["encounters_path"]):
                self.import_encounters_json(self.run_config["encounters_path"])

        # update honey table encounters
        if self.GameData.game_name == "Platinum":
            self.GameData.honey_mons = [
                mon
                for mon in self.GameData.honey_mons
                if mon in self.GameData.encounter_table
            ]

        # notes for encounter order on final spreadsheet
        self.notes = {route: {} for route in list(self.GameData.encounter_table.index)}

        if self.args.route:
            self.route()

    def import_encounters_json(self, encounters_path):
        with open(encounters_path, "r") as f:
            assign_me = json.loads(f.read())

        routes = list(route for route in assign_me.keys() if assign_me[route] != "")
        failed_routes = list(
            route for route in assign_me.keys() if assign_me[route] == ""
        )
        encounters = [
            list(self.GameData.encounter_table.filter(like=encounter, axis=1).columns)[
                0
            ]
            for encounter in assign_me.values()
            if encounter != ""
            and list(
                self.GameData.encounter_table.filter(like=encounter, axis=1).columns
            )
            != []
        ]

        assign_me = {routes[i]: encounters[i] for i in range(len(routes))}
        routes = routes + failed_routes

        group_data = GroupData(
            assign_me=assign_me, routes=routes, encounters=encounters
        )
        self.update(
            group_data=group_data, df=self.GameData.encounter_table, is_json=True
        )

    def init_parser(self):
        parser = argparse.ArgumentParser()
        group = parser.add_mutually_exclusive_group(required=True)
        group.add_argument(
            "--test",
            "-t",
            required=False,
            help="select to test manual encounter routing",
            action="store_true",
        )
        group.add_argument(
            "--route",
            "-r",
            required=False,
            help="select to use the routing algorithm",
            action="store_true",
        )
        parser.add_argument(
            "--config", "-c", required=True, help=".yaml file for configuring run data"
        )
        parser.add_argument(
            "--cloud",
            "-cl",
            required=False,
            help="toggle loading encounter data from google sheet",
            action="store_true",
        )
        return parser

    def route(self):
        methods = {self.assign_one_to_one: True, self.check_duplicates: True}

        while True in methods.values():
            for method in methods.keys():
                methods[method] = method()

        # ask user about unique encounters before finishing?
        self.print_progress(
            pd.DataFrame.from_dict({"Encounter": self.assigned_encounters}), "*=*=*"
        )
        self.export_table()

    def assign_one_to_one(self):
        assigned = False
        working_df = self.encounter_tables[-1].copy(deep=True)

        only_one_bool = working_df.sum(axis=1) == 1

        while only_one_bool.any():
            assigned = True
            # get routes with only one eligible encounter (pd.DataFrame)
            # filter out the irrelevant encounters
            only_one_options = (
                working_df[only_one_bool].replace(0, pd.NA).dropna(axis=1, how="all")
            )

            # melt to pd.DataFrame (index: routes, column: encounter)
            only_one_melted = pd.melt(
                only_one_options, ignore_index=False, var_name="Encounter"
            ).dropna()[["Encounter"]]
            only_one_dict = only_one_melted.to_dict("split")
            group_data = GroupData(
                routes=only_one_dict["index"],
                encounters=list(map(lambda x: x[0], only_one_dict["data"])),
                assign_me={
                    only_one_dict["index"][i]: only_one_dict["data"][i][0]
                    for i in range(len(only_one_dict["index"]))
                },
            )

            working_df = self.update(group_data, working_df)

            only_one_bool = working_df.sum(axis=1) == 1

        return assigned

    def check_duplicates(self):
        working_df = self.encounter_tables[-1].copy(deep=True)
        relevant_df = working_df[working_df.duplicated(keep=False)]
        groups = relevant_df.groupby(by=list(relevant_df.columns), sort=False).groups

        # collect eligible group data and ask the player which group to assign
        groups_data = {}
        i = 0
        for group in groups:
            encounters = [
                relevant_df.columns[i]
                for i in range(len(relevant_df.columns))
                if group[i] == 1
            ]
            flag_honey = (
                True
                if set(encounters).issuperset(
                    set(
                        [
                            mon
                            for mon in self.GameData.honey_mons
                            if mon in working_df.columns
                        ]
                    )
                )
                else False
            )

            group_data = GroupData()
            # if there are more routes than encounters, each encounter can be assigned
            if len(groups[group].values) >= np.nansum(group):
                group_data.encounters = encounters
                group_data.routes = list(groups[group].values)
                group_data.assign_me = {
                    route: " or ".join(encounters) for route in group_data.routes
                }
                group_data.flags["Honey"] = flag_honey
                groups_data[i] = group_data
                i += 1

        # check if user would like to update any group
        if len(groups_data.keys()) == 1:
            self.update(groups_data[0], working_df)
            return True
        elif groups_data != {}:
            print(groups_data)
            group_data = groups_data.get(
                int(input("Which group would you like to assign?\n> "))
            )
            if group_data:
                self.update(group_data, working_df)
                return True
        else:
            return False

    def update(self, group_data, df, is_json=False):
        # if honey location, get all remaining honey encounters and update
        if group_data.flags["Honey"]:
            # update honey table encounters
            self.GameData.honey_mons = [
                mon for mon in self.GameData.honey_mons if mon in df
            ]
            remaining_honey_routes = [
                index
                for index, row in df.iterrows()
                if sum(row[self.GameData.honey_mons])
                == len(row[self.GameData.honey_mons])
                and sum(row[self.GameData.honey_mons]) == row.sum()
                and index not in group_data.routes
            ]
            flag_group_data = GroupData(remaining_honey_routes, self.GameData.honey_mons)
            flag_group_data.assign_me = {
                route: " or ".join(flag_group_data.encounters)
                for route in flag_group_data.routes
            }
            self.update(flag_group_data, df)

        # update the assigned encounter dict
        self.assigned_encounters.update(group_data.assign_me)

        # make note of routes with encounters that have been assigned
        if not is_json:
            for mon in group_data.encounters:
                mon_routes = [
                    route
                    for route in df[mon][df[mon] == 1].index
                    if route not in group_data.routes
                ]
                for route in mon_routes:
                    pass_n = len(self.encounter_tables)
                    curr_notes = self.notes[route].get(pass_n)
                    if curr_notes:
                        curr_notes.add(mon)
                    else:
                        self.notes[route][pass_n] = set([mon])

        # remove filtered items from df
        df = df.loc[
            ~df.index.isin(group_data.routes), ~df.columns.isin(group_data.encounters)
        ].dropna(axis=0, how="all")

        # remove encounters that are no longer accessible
        self.check_available_encounters(df)

        # add the updated dataframe as a slice to the results table
        self.encounter_tables.append(df)
        if is_json:
            self.encounter_tables.pop(0)

        # print encounters added
        new_encounters = pd.DataFrame.from_dict(
            group_data.assign_me, orient="index", columns=["Encounter"]
        )
        new_encounters["Pass"] = len(self.encounter_tables) - 1
        new_encounters.index.name = "Route"
        new_encounters.set_index(["Pass", new_encounters.index], inplace=True)
        self.assigned_encounters_slice.append(new_encounters)
        self.print_progress(new_encounters)

        # return df because assign_one_to_one needs the working_df updated to continue making passes
        # alternatively could make it loop outside of the method?
        return df

    def check_available_encounters(self, df):
        lost_encounters = df.loc[:, df.sum(axis=0) == 0]
        df.dropna(axis=1, how="all", inplace=True)
        if len(lost_encounters.columns) != 0:
            print("WARNING: {} ARE LOST".format(", ".join(lost_encounters.columns)))

    def print_progress(self, df, pattern="="):
        print(
            pattern * (50 // len(pattern)) + pattern[: 50 % len(pattern)],
            "Update {}:".format(len(self.encounter_tables) - 1)
            if pattern == "="
            else "Final Encounters:",
            df,
            pattern * (50 // len(pattern)) + pattern[: 50 % len(pattern)],
            "\n",
            sep="\n",
        )

    def export_table(self):
        # REMOVE HASHES FROM COLORS TO MAKE THEM USABLE
        green_fill = openpyxl.styles.PatternFill(bgColor="c6efce", fill_type="solid")
        green_font = openpyxl.styles.Font(color="006100")
        yellow_fill = openpyxl.styles.PatternFill(bgColor="FFEB9C", fill_type="solid")
        yellow_font = openpyxl.styles.Font(color="9C5700")
        yellowdxf = openpyxl.styles.differential.DifferentialStyle(
            font=yellow_font, fill=yellow_fill
        )

        with pd.ExcelWriter(self.run_config["output_sheet_path"]) as writer:
            # write each pass' slice as a worksheet
            for i in range(len(self.encounter_tables)):
                self.encounter_tables[i].to_excel(
                    writer,
                    sheet_name="EncounterTable{}".format(str(i)),
                    freeze_panes=(1, 1),
                )

                worksheet = writer.sheets["EncounterTable{}".format(str(i))]

                sheet_data = list(worksheet.columns)
                # the range of the table body ignoring header row and column
                c_range = "{}:{}".format(
                    sheet_data[1][1].coordinate, sheet_data[-1][-1].coordinate
                )

                # fills the background color
                worksheet.conditional_formatting.add(
                    c_range,
                    openpyxl.formatting.rule.CellIsRule(
                        operator="equal",
                        formula=[1],
                        stopIfTrue=False,
                        fill=green_fill,
                        font=green_font,
                    ),
                )

            # write the historical pass table as the final worksheet
            pd.concat(self.assigned_encounters_slice).to_excel(
                writer, sheet_name="EncounterList"
            )
            worksheet = writer.sheets["EncounterList"]
            # the range of the relevant column
            c_range = "{}:{}".format(
                worksheet["C"][0].coordinate, worksheet["C"][-1].coordinate
            )
            worksheet.conditional_formatting.add(
                c_range,
                openpyxl.formatting.rule.Rule(type="duplicateValues", dxf=yellowdxf),
            )

            # add comments to each sheet to show what pokemon need to be encountered to create that instance of encounter routing
            for sheet in writer.sheets:
                worksheet = writer.sheets[sheet]

                if not sheet.isalpha():
                    i = i = int("".join([char for char in sheet if char.isnumeric()]))

                if "List" in sheet:
                    min_c, max_c = 2, 2
                else:
                    min_c, max_c = 1, 1

                for col in worksheet.iter_cols(min_row=2, min_col=min_c, max_col=max_c):
                    for cell in col:
                        route_data = self.notes[cell.value]
                        message = "\n".join(
                            [
                                "{}: {}".format(p, ", ".join(route_data[p]))
                                for p in range(1, i + 1)
                                if route_data.get(p)
                            ]
                        )
                        if message != "":
                            cell.comment = openpyxl.comments.Comment(
                                message, "openpyxl"
                            )
        return


if __name__ == "__main__":
    NRotMEncounterRouting()
